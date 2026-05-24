"""
CONDO-SAFE24 — Sistema SaaS de Segurança para Condomínios
Versão 3.0 — Arquitetura de Privacidade por Design (LGPD)

Separação total: identidade cifrada no vault / ocorrência no banco operacional.
Central NÃO vê nome do morador. Admin autorizado usa /api/reveal (auditado).

SpyNet Tecnologia Forense | CNPJ 64.000.808/0001-51 | spynetintelligence@proton.me
"""

from flask import (
    Flask, render_template, send_file, jsonify,
    request, redirect, url_for, session
)
from flask_wtf.csrf import CSRFProtect
import os, json, time, secrets, logging
from pathlib import Path
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from collections import defaultdict
from functools import wraps
import hmac

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas

app = Flask(__name__)
csrf = CSRFProtect(app)

SECRET_KEY = os.environ.get('SECRET_KEY', '').strip()
if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)
    logging.warning("SECRET_KEY não definida. Configure uma chave fixa no Render ENV!")

app.secret_key = SECRET_KEY
app.config.update(
    JSONIFY_PRETTYPRINT_REGULAR=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.environ.get('FLASK_ENV') == 'production',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    WTF_CSRF_ENABLED=True,
    WTF_CSRF_TIME_LIMIT=3600,
)

TZ = ZoneInfo(os.environ.get('APP_TZ', 'America/Sao_Paulo'))
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger('condosafe')

# ── Estado em memória (cache — SQLite é a fonte de verdade) ───────────────────
alertas        = []
sistema_status = {'sirene_ativa': False, 'mutado': False, 'ultima_atualizacao': None}

_rate_store     = defaultdict(list)
_login_failures = defaultdict(list)

# ── Código de revelação de identidade ─────────────────────────────────────────
REVEAL_SECRET = os.environ.get('REVEAL_SECRET', '').strip()


# ════════════════════════════════════════════════════════════════════════════
#  HELPERS DE IP, RATE LIMIT, SANITIZAÇÃO
# ════════════════════════════════════════════════════════════════════════════

def _get_ip():
    return (
        request.headers.get('X-Forwarded-For', '').split(',')[0].strip()
        or request.headers.get('X-Real-IP', '')
        or request.remote_addr or '0.0.0.0'
    )

def _rate_limit(key, max_calls=60, window=60):
    now = time.time()
    _rate_store[key] = [t for t in _rate_store[key] if now - t < window]
    if len(_rate_store[key]) >= max_calls:
        return True
    _rate_store[key].append(now)
    return False

def _login_rate_limit(ip):
    now = time.time()
    _login_failures[ip] = [t for t in _login_failures[ip] if now - t < 300]
    return len(_login_failures[ip]) >= 10

def _record_login_failure(ip):
    _login_failures[ip].append(time.time())

def _sanitize(v, maxlen=200):
    if not isinstance(v, str):
        v = str(v) if v is not None else ''
    for ch in ['<', '>', '"', "'", '\\', '\x00']:
        v = v.replace(ch, '')
    return v.strip()[:maxlen]


# ════════════════════════════════════════════════════════════════════════════
#  HEADERS DE SEGURANÇA
# ════════════════════════════════════════════════════════════════════════════

@app.after_request
def set_security_headers(resp):
    resp.headers['X-Content-Type-Options']  = 'nosniff'
    resp.headers['X-Frame-Options']         = 'SAMEORIGIN'
    resp.headers['X-XSS-Protection']        = '1; mode=block'
    resp.headers['Referrer-Policy']         = 'strict-origin-when-cross-origin'
    resp.headers['Permissions-Policy']      = 'geolocation=(self)'
    resp.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    resp.headers['Content-Security-Policy']   = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "connect-src 'self'; "
        "img-src 'self' data:; "
        "font-src 'self'"
    )
    return resp


# ════════════════════════════════════════════════════════════════════════════
#  GESTÃO DE CHAVES (por condomínio)
# ════════════════════════════════════════════════════════════════════════════

KEYS_FILE = Path(os.environ.get('CONDO_KEYS_FILE', 'data/keys.json'))

def _ensure_data_dir():
    KEYS_FILE.parent.mkdir(parents=True, exist_ok=True)

def _load_keys():
    out = {}
    env_json = (os.environ.get('CONDO_KEYS_JSON') or '').strip()
    if env_json:
        try:
            data = json.loads(env_json)
            if isinstance(data, dict):
                for k, v in data.items():
                    out[str(k)] = {
                        'name':   str(v.get('name', 'Condomínio') if isinstance(v, dict) else v).strip() or 'Condomínio',
                        'active': bool(v.get('active', True) if isinstance(v, dict) else True)
                    }
        except Exception:
            pass
    try:
        if KEYS_FILE.exists():
            data = json.loads(KEYS_FILE.read_text(encoding='utf-8'))
            if isinstance(data, dict):
                for k, v in data.items():
                    out[str(k)] = {
                        'name':   str(v.get('name', 'Condomínio') if isinstance(v, dict) else v).strip() or 'Condomínio',
                        'active': bool(v.get('active', True) if isinstance(v, dict) else True)
                    }
    except Exception:
        pass
    if not out:
        _ensure_data_dir()
        demo = {'DEMO-1234': {'name': 'Condomínio Demo', 'active': True}}
        try:
            KEYS_FILE.write_text(json.dumps(demo, ensure_ascii=False, indent=2), encoding='utf-8')
        except Exception:
            pass
        return demo
    return out

def _save_keys(keys):
    _ensure_data_dir()
    KEYS_FILE.write_text(json.dumps(keys, ensure_ascii=False, indent=2), encoding='utf-8')

def _is_valid_client_key(key):
    if not key:
        return False
    info = _load_keys().get(key)
    return bool(info and info.get('active', True))

def _client_info(key):
    return _load_keys().get(key) or {}

def _require_client_key():
    key = (
        (request.args.get('key') or '').strip()
        or (request.headers.get('X-CLIENT-KEY') or '').strip()
        or ((request.get_json(silent=True) or {}).get('key') or '').strip()
        or (session.get('client_key') or '').strip()
    )
    return key if _is_valid_client_key(key) else None


# ════════════════════════════════════════════════════════════════════════════
#  DECORATORS DE AUTENTICAÇÃO
# ════════════════════════════════════════════════════════════════════════════

def _require_central_or_admin():
    ec = (os.environ.get('CENTRAL_PASSWORD') or '').strip()
    ea = (os.environ.get('ADMIN_PASSWORD') or '').strip()
    return ((not ec) or bool(session.get('central_auth'))) or \
           ((not ea) or bool(session.get('admin_auth')))

def require_central(f):
    @wraps(f)
    def decorated(*a, **kw):
        if not _require_central_or_admin():
            if request.is_json:
                return jsonify({'ok': False, 'error': 'Não autorizado.'}), 401
            return redirect(url_for('login_central'))
        return f(*a, **kw)
    return decorated

def require_admin(f):
    @wraps(f)
    def decorated(*a, **kw):
        expected = (os.environ.get('ADMIN_PASSWORD') or '').strip()
        if expected and not session.get('admin_auth'):
            if request.is_json:
                return jsonify({'ok': False, 'error': 'Não autorizado.'}), 401
            return redirect(url_for('login_admin'))
        return f(*a, **kw)
    return decorated


# ════════════════════════════════════════════════════════════════════════════
#  ROTAS DE PÁGINAS
# ════════════════════════════════════════════════════════════════════════════

@app.route('/')
def home():
    return redirect(url_for('landing'))

@app.route('/painel')
def painel_home():
    return render_template('home.html')

@app.route('/sos')
def sos():
    key_q = (request.args.get('key') or '').strip()
    if key_q:
        if not _is_valid_client_key(key_q):
            return render_template('forbidden.html',
                msg='Chave inválida ou desativada. Solicite o link correto ao administrador.'), 403
        session['client_key']  = key_q
        session['client_name'] = _client_info(key_q).get('name', 'Condomínio')
        return redirect(url_for('sos'))
    key = (session.get('client_key') or '').strip()
    if not _is_valid_client_key(key):
        return render_template('forbidden.html',
            msg='Acesso restrito. Use o link oficial com a chave do seu condomínio.'), 403
    return render_template('sos.html', client_name=session.get('client_name', 'Condomínio'))

@app.route('/professor')
def professor():
    return redirect(url_for('sos', **request.args))

@app.route('/central')
@require_central
def central():
    return render_template('central.html')

@app.route('/painel_publico')
@require_central
def painel_publico():
    return render_template('painel_publico.html')

@app.route('/admin')
@require_admin
def admin():
    return render_template('admin.html')

@app.route('/analytics')
@require_central
def analytics_page():
    return render_template('analytics.html')

@app.route('/landing')
@app.route('/planos')
@app.route('/vendas')
def landing():
    return render_template('landing.html')

@app.route('/demo')
def demo():
    return render_template('demo.html')


# ════════════════════════════════════════════════════════════════════════════
#  ROTAS DE LOGIN / LOGOUT
# ════════════════════════════════════════════════════════════════════════════

@app.route('/login_central', methods=['GET', 'POST'])
def login_central():
    ip = _get_ip()
    if request.method == 'POST':
        if _login_rate_limit(ip):
            return render_template('login_central.html', erro='Muitas tentativas. Aguarde alguns minutos.'), 429
        senha    = (request.form.get('senha') or request.form.get('password') or '').strip()
        expected = (os.environ.get('CENTRAL_PASSWORD') or '').strip()
        if expected and not hmac.compare_digest(senha.encode(), expected.encode()):
            return render_template('login_central.html', erro='Senha incorreta.')
        session.permanent = True
        session['central_auth'] = True
        return redirect(url_for('central'))
    return render_template('login_central.html', erro=None)

@app.route('/login_admin', methods=['GET', 'POST'])
def login_admin():
    ip = _get_ip()
    if request.method == 'POST':
        if _login_rate_limit(ip):
            return render_template('login_admin.html', erro='Muitas tentativas. Aguarde alguns minutos.'), 429
        senha    = (request.form.get('senha') or request.form.get('password') or '').strip()
        expected = (os.environ.get('ADMIN_PASSWORD') or '').strip()
        if expected and not hmac.compare_digest(senha.encode(), expected.encode()):
            _record_login_failure(ip)
            log.warning(f"Login Admin falhou — IP: {ip}")
            return render_template('login_admin.html', erro='Senha incorreta.')
        session.permanent = True
        session['admin_auth'] = True
        return redirect(url_for('admin'))
    return render_template('login_admin.html', erro=None)

@app.route('/logout_central')
def logout_central():
    session.pop('central_auth', None)
    return redirect(url_for('login_central'))

@app.route('/logout_admin')
def logout_admin():
    session.pop('admin_auth', None)
    return redirect(url_for('login_admin'))

@app.route('/play-alarm')
@app.route('/tocar_sirene')
def play_alarm():
    try:
        return send_file('static/siren.wav.mp3')
    except FileNotFoundError:
        return 'Arquivo de áudio não encontrado', 404


# ════════════════════════════════════════════════════════════════════════════
#  API — RECEBER ALERTA (v3.0 — separa identidade de ocorrência)
# ════════════════════════════════════════════════════════════════════════════

@csrf.exempt
@app.route('/api/alert', methods=['POST'])
def receber_alerta():
    ip = _get_ip()
    if _rate_limit(f'alert:{ip}', max_calls=20, window=60):
        return jsonify({'ok': False, 'error': 'Muitas requisições.'}), 429

    client_key = _require_client_key()
    if not client_key:
        log.warning(f"Alerta sem chave válida — IP: {ip}")
        return jsonify({'ok': False, 'error': 'Acesso negado (chave inválida).'}), 403

    try:
        data = request.get_json() or {}

        # Dados de identidade → vão para o vault (cifrados)
        caller      = _sanitize(data.get('caller') or data.get('teacher') or 'Morador')
        contact     = _sanitize(data.get('contact') or '')
        unidade     = _sanitize(data.get('unidade') or '')

        # Dados operacionais → vão para alertas (sem identidade)
        area        = _sanitize(data.get('area') or data.get('location') or data.get('room') or '')
        occ_type    = _sanitize(data.get('type') or data.get('occ_type') or 'Pânico')
        description = _sanitize(data.get('description') or 'Sem descrição', maxlen=500)
        prioridade  = int(data.get('prioridade', 2))

        lat = data.get('lat'); lng = data.get('lng'); accuracy = data.get('accuracy')
        maps_url = None
        if lat is not None and lng is not None:
            try:
                lat = float(lat); lng = float(lng)
                maps_url = f"https://www.google.com/maps?q={lat},{lng}"
            except (ValueError, TypeError):
                lat = lng = None

        now = datetime.now(TZ)

        # Monta pacote completo para db.py processar
        pacote = {
            # identidade (vai para vault cifrado)
            'caller':       caller,
            'contact':      contact,
            'unidade':      unidade,
            # ocorrência (vai para alertas sem dados pessoais)
            'area':         area,
            'type':         occ_type,
            'prioridade':   prioridade,
            'description':  description,
            'lat':          lat, 'lng': lng, 'accuracy': accuracy,
            'maps_url':     maps_url,
            'client_key':   client_key,
            'client_name':  _client_info(client_key).get('name', 'Condomínio'),
            'ip':           ip,
            'timestamp':    now.strftime('%d/%m/%Y %H:%M:%S'),
        }

        # Salva via db.py (separação automática)
        try:
            resultado = salvar_alerta(pacote)
            token_id  = resultado['token_id']
            alert_id  = resultado['alert_id']
        except Exception as e:
            log.warning(f"SQLite indisponível, usando memória: {e}")
            token_id = gerar_token_id()
            alert_id = len(alertas) + 1

        # Cache em memória (para polling rápido da central)
        novo = {
            'id':          alert_id,
            'token_id':    token_id,
            'area':        area,
            'type':        occ_type,
            'prioridade':  prioridade,
            'description': description,
            'timestamp':   now.strftime('%d/%m/%Y %H:%M:%S'),
            'resolved':    False,
            'ts':          now.strftime('%H:%M:%S'),
            'lat':         lat, 'lng': lng, 'accuracy': accuracy,
            'maps_url':    maps_url,
            'client_key':  client_key,
            'client_name': _client_info(client_key).get('name', 'Condomínio'),
            # ⚠️ JAMAIS incluir caller, contact ou unidade aqui
        }
        alertas.insert(0, novo)
        if len(alertas) > 500:
            alertas[:] = alertas[:500]

        sistema_status['sirene_ativa']       = True
        sistema_status['ultima_atualizacao'] = now.isoformat()

        log.info(f"ALERTA #{alert_id} token={token_id} tipo={occ_type} condo={client_key}")

        # Notificação WhatsApp
        _enviar_whatsapp(
            f"🚨 *CONDO-SAFE24 — ALERTA #{alert_id}*\n"
            f"🏢 *{_client_info(client_key).get('name', 'Condomínio')}*\n"
            f"⚠️ Tipo: {occ_type}\n"
            f"📍 Área: {area or 'Não informada'}\n"
            f"🕐 Hora: {now.strftime('%H:%M:%S')}\n"
            f"🔑 Token: {token_id}\n"
            + (f"📌 GPS: {maps_url}" if maps_url else "")
        )

        return jsonify({'ok': True, 'alert_id': alert_id, 'token_id': token_id})

    except Exception:
        log.exception("Erro em /api/alert")
        return jsonify({'ok': False, 'error': 'Erro interno.'}), 500


# ════════════════════════════════════════════════════════════════════════════
#  API — STATUS DA CENTRAL (sem dados pessoais)
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/status', methods=['GET'])
@require_central
def status_sistema():
    ip = _get_ip()
    if _rate_limit(f'status:{ip}', max_calls=120, window=60):
        return jsonify({'ok': False, 'error': 'Limite excedido.'}), 429

    try:
        dados = listar_alertas()
    except Exception:
        dados = alertas

    ativos = [a for a in dados if not a.get('resolved')]
    return jsonify({
        'ok':           True,
        'siren':        sistema_status['sirene_ativa'],
        'muted':        sistema_status['mutado'],
        'alerts':       dados,          # ← sem nome, sem contato, sem unidade
        'active_alerts': len(ativos),
        'total_alerts':  len(dados),
        'last_update':   sistema_status['ultima_atualizacao'],
    })


# ════════════════════════════════════════════════════════════════════════════
#  API — REVELAR IDENTIDADE (admin + REVEAL_SECRET, auditado)
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/reveal', methods=['POST'])
@require_admin
def revelar_identidade_api():
    ip = _get_ip()
    if _rate_limit(f'reveal:{ip}', max_calls=10, window=60):
        return jsonify({'ok': False, 'error': 'Muitas requisições de revelação.'}), 429

    data       = request.get_json() or {}
    token_id   = _sanitize(data.get('token_id') or '')
    codigo     = (data.get('reveal_secret') or '').strip()

    # Verifica código de revelação (se configurado)
    if REVEAL_SECRET and codigo != REVEAL_SECRET:
        log.warning(f"Tentativa de revelar identidade com código errado — IP: {ip}")
        audit('REVEAL_DENIED', token_id=token_id, operator='admin', ip=ip,
              details='Código REVEAL_SECRET incorreto')
        return jsonify({'ok': False, 'error': 'Código de autorização inválido.'}), 403

    if not token_id:
        return jsonify({'ok': False, 'error': 'token_id obrigatório.'}), 400

    try:
        resultado = revelar_identidade(
            token_id  = token_id,
            operator  = 'admin',
            ip        = ip
        )
        if not resultado.get('encontrado'):
            return jsonify({'ok': False, 'error': 'Token não encontrado ou dados expirados.'}), 404

        return jsonify({
            'ok':       True,
            'dados':    resultado,
            'ttl':      300,       # cliente deve apagar da tela em 5 minutos
            'aviso':    'Dados visíveis por 5 minutos. Acesso registrado em auditoria.'
        })
    except Exception:
        log.exception("Erro em /api/reveal")
        return jsonify({'ok': False, 'error': 'Erro interno.'}), 500


# ════════════════════════════════════════════════════════════════════════════
#  API — SIRENE, RESOLVER, LIMPAR
# ════════════════════════════════════════════════════════════════════════════

@csrf.exempt
@app.route('/api/siren', methods=['POST'])
@require_central
def controlar_sirene():
    data   = request.get_json() or {}
    action = data.get('action')
    if action == 'on':
        sistema_status['sirene_ativa'] = True;  sistema_status['mutado'] = False
    elif action == 'off':
        sistema_status['sirene_ativa'] = False; sistema_status['mutado'] = False
    elif action == 'mute':
        sistema_status['mutado'] = True
    else:
        return jsonify({'ok': False, 'error': 'Ação inválida.'}), 400
    sistema_status['ultima_atualizacao'] = datetime.now(TZ).isoformat()
    return jsonify({'ok': True, 'siren': sistema_status['sirene_ativa'],
                    'muted': sistema_status['mutado']})

@csrf.exempt
@app.route('/api/resolve', methods=['POST'])
@require_central
def resolver_alerta():
    data     = request.get_json() or {}
    alert_id = data.get('id')
    token_id = data.get('token_id')

    for a in alertas:
        if alert_id and str(a.get('id')) == str(alert_id) and not a['resolved']:
            a['resolved'] = True; break
        elif not alert_id and not a['resolved']:
            a['resolved'] = True; break

    try:
        resolver_alerta_db(alert_id=alert_id, token_id=token_id)
    except Exception as e:
        log.warning(f"Erro ao resolver no SQLite: {e}")

    if not any(a for a in alertas if not a['resolved']):
        sistema_status['sirene_ativa'] = False

    sistema_status['ultima_atualizacao'] = datetime.now(TZ).isoformat()
    return jsonify({'ok': True})

@csrf.exempt
@app.route('/api/clear', methods=['POST'])
@require_central
def limpar_alertas():
    alertas.clear()
    sistema_status['sirene_ativa']       = False
    sistema_status['ultima_atualizacao'] = datetime.now(TZ).isoformat()
    try:
        limpar_alertas_db()
    except Exception as e:
        log.warning(f"Erro ao limpar SQLite: {e}")
    log.info(f"Alertas limpos por {_get_ip()}")
    audit('CLEAR_ALERTS', operator='central', ip=_get_ip())
    return jsonify({'ok': True})

# Rota legada (compatibilidade)
@csrf.exempt
@app.route('/acionar_alerta', methods=['POST'])
def acionar_alerta():
    client_key = _require_client_key()
    if not client_key:
        return jsonify({'success': False, 'message': 'Acesso negado.'}), 403
    now  = datetime.now(TZ)
    novo = {
        'id': len(alertas) + 1,
        'token_id':   gerar_token_id(),
        'area':       'Local não informado',
        'type':       'Pânico',
        'prioridade': 2,
        'description':'Alerta de pânico acionado',
        'timestamp':  now.strftime('%d/%m/%Y %H:%M:%S'),
        'resolved':   False,
        'ts':         now.strftime('%H:%M:%S'),
        'client_key': client_key,
        'client_name': _client_info(client_key).get('name', 'Condomínio')
    }
    alertas.append(novo)
    sistema_status['sirene_ativa'] = True
    return jsonify({'success': True, 'message': 'Alerta acionado!', 'alerta': novo})


# ════════════════════════════════════════════════════════════════════════════
#  API — ANALYTICS
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/analytics')
@require_central
def analytics():
    try:
        try:
            s = stats_alertas()
        except Exception:
            from collections import Counter
            total  = len(alertas)
            ativos = sum(1 for a in alertas if not a['resolved'])
            tipos  = Counter(a.get('type', '') for a in alertas)
            s = {
                'total': total, 'ativos': ativos, 'resolvidos': total - ativos,
                'por_tipo': [{'type': k, 'n': v} for k, v in tipos.most_common()],
                'por_hora': [],
            }
        return jsonify({'ok': True, 'stats': s})
    except Exception:
        log.exception("Erro em /api/analytics")
        return jsonify({'ok': False, 'error': 'Erro interno.'}), 500


# ════════════════════════════════════════════════════════════════════════════
#  API — ADMIN: KEYS
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/keys', methods=['GET'])
@require_admin
def list_keys():
    return jsonify({'ok': True, 'keys': _load_keys()})

@app.route('/api/keys', methods=['POST'])
@require_admin
def create_key():
    data     = request.get_json() or {}
    name     = _sanitize(data.get('name') or 'Condomínio', maxlen=80)
    new_key  = 'CONDO-' + secrets.token_urlsafe(10).replace('-', '').replace('_', '')[:12].upper()
    keys     = _load_keys()
    keys[new_key] = {'name': name, 'active': True}
    try:
        _save_keys(keys)
    except Exception as e:
        log.error(f"Erro ao salvar chave: {e}")
    audit('CREATE_KEY', operator='admin', ip=_get_ip(), details=f"Chave {new_key} para {name}")
    return jsonify({'ok': True, 'key': new_key, 'name': name})

@app.route('/api/keys/toggle', methods=['POST'])
@require_admin
def toggle_key():
    data   = request.get_json() or {}
    key    = (data.get('key') or '').strip()
    active = bool(data.get('active', True))
    keys   = _load_keys()
    if key not in keys:
        return jsonify({'ok': False, 'error': 'Chave não encontrada'}), 404
    info   = keys.get(key) or {}
    keys[key] = {'name': (info.get('name') if isinstance(info, dict) else str(info)), 'active': active}
    try:
        _save_keys(keys)
    except Exception as e:
        log.error(f"Erro ao toggle: {e}")
    audit('TOGGLE_KEY', operator='admin', ip=_get_ip(), details=f"Chave {key} → active={active}")
    return jsonify({'ok': True})

@app.route('/api/keys/delete', methods=['POST'])
@require_admin
def delete_key():
    data = request.get_json() or {}
    key  = (data.get('key') or '').strip()
    keys = _load_keys()
    if key not in keys:
        return jsonify({'ok': False, 'error': 'Chave não encontrada'}), 404
    del keys[key]
    try:
        _save_keys(keys)
    except Exception as e:
        log.error(f"Erro ao deletar: {e}")
    audit('DELETE_KEY', operator='admin', ip=_get_ip(), details=f"Chave {key} removida")
    return jsonify({'ok': True})


# ════════════════════════════════════════════════════════════════════════════
#  API — ADMIN: AUDITORIA
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/audit')
@require_admin
def audit_log_api():
    try:
        registros = listar_audit(limit=200)
        return jsonify({'ok': True, 'logs': registros})
    except Exception:
        return jsonify({'ok': False, 'error': 'Erro ao carregar auditoria.'}), 500


# ════════════════════════════════════════════════════════════════════════════
#  API — ADMIN: EXPIRAR VAULT MANUALMENTE
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/vault/expire', methods=['POST'])
@require_admin
def vault_expire():
    try:
        removed = expirar_vault()
        return jsonify({'ok': True, 'removed': removed})
    except Exception:
        return jsonify({'ok': False, 'error': 'Erro ao expirar vault.'}), 500


# ════════════════════════════════════════════════════════════════════════════
#  HEALTH CHECK
# ════════════════════════════════════════════════════════════════════════════

@app.route('/health')
def health_check():
    return jsonify({
        'status':    'healthy',
        'version':   '3.0-PRIVACY',
        'timestamp': datetime.now(TZ).isoformat(),
        'alerts':    len(alertas),
        'active':    sum(1 for a in alertas if not a.get('resolved')),
    })


# ════════════════════════════════════════════════════════════════════════════
#  PDF REPORT (sem dados pessoais — apenas tokens e ocorrências)
# ════════════════════════════════════════════════════════════════════════════

@app.route('/report.pdf')
@require_central
def report_pdf():
    buf  = BytesIO()
    c    = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    now  = datetime.now(TZ)

    try:
        dados = listar_alertas()
    except Exception:
        dados = alertas

    total     = len(dados)
    ativos    = sum(1 for a in dados if not a.get('resolved'))
    resolvidos = total - ativos
    taxa      = int(resolvidos / total * 100) if total > 0 else 0

    def draw_frame(pg=1):
        c.setFillColorRGB(0.04, 0.07, 0.15)
        c.rect(0, H - 75, W, 75, fill=1, stroke=0)
        c.setFillColorRGB(0, 0.6, 1); c.setFont("Helvetica-Bold", 18)
        c.drawString(40, H - 38, "CONDO-SAFE24 — Relatório Operacional (sem dados pessoais)")
        c.setFillColorRGB(0.75, 0.8, 0.95); c.setFont("Helvetica", 9)
        c.drawString(40, H - 54, "SpyNet Tecnologia Forense | LGPD — Identidades protegidas no vault cifrado")
        c.setFillColorRGB(0, 0.8, 0.6); c.setFont("Helvetica-Bold", 9)
        c.drawRightString(W - 40, H - 38, f"Gerado: {now.strftime('%d/%m/%Y %H:%M:%S')}")
        c.setFillColorRGB(0.5, 0.5, 0.7); c.setFont("Helvetica", 8)
        c.drawRightString(W - 40, H - 52, f"Página {pg}")
        c.setStrokeColorRGB(0, 0.5, 0.9); c.setLineWidth(1.5); c.line(0, H - 77, W, H - 77)
        c.setStrokeColorRGB(0.2, 0.2, 0.3); c.setLineWidth(0.5); c.line(40, 40, W - 40, 40)
        c.setFillColorRGB(0.4, 0.4, 0.5); c.setFont("Helvetica", 7.5)
        c.drawString(40, 28, "CONDO-SAFE24 © SpyNet Tecnologia Forense | CNPJ 64.000.808/0001-51")
        c.drawRightString(W - 40, 28, "Documento confidencial — sem dados pessoais")

    draw_frame(1)
    y = H - 100

    boxes = [
        ("Total", str(total), (0.05, 0.09, 0.2)),
        ("Ativos", str(ativos), (0.45, 0.07, 0.07)),
        ("Resolvidos", str(resolvidos), (0.04, 0.28, 0.18)),
        (f"Taxa {taxa}%", "Resolução", (0.04, 0.18, 0.38)),
    ]
    bw = (W - 80 - 30) / 4
    for i, (l, v, bg) in enumerate(boxes):
        bx = 40 + i * (bw + 10)
        c.setFillColorRGB(*bg); c.roundRect(bx, y - 52, bw, 52, 5, fill=1, stroke=0)
        c.setFillColorRGB(0.6, 0.7, 0.9); c.setFont("Helvetica", 7.5)
        c.drawCentredString(bx + bw / 2, y - 13, l)
        c.setFillColorRGB(1, 1, 1); c.setFont("Helvetica-Bold", 18 if len(v) < 5 else 13)
        c.drawCentredString(bx + bw / 2, y - 36, v)
    y -= 72

    c.setFillColorRGB(0.85, 0.88, 0.97); c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Histórico de Ocorrências (tokens anônimos)"); y -= 18

    cols = [("#", 25), ("Token", 90), ("Data/Hora", 100), ("Tipo", 90),
            ("Área", 80), ("GPS", 100), ("Status", 65)]
    c.setFillColorRGB(0.04, 0.07, 0.22)
    c.rect(40, y - 14, W - 80, 16, fill=1, stroke=0)
    c.setFillColorRGB(0, 0.7, 1); c.setFont("Helvetica-Bold", 7.5)
    xc = 45
    for cn, cw in cols:
        c.drawString(xc, y - 9, cn); xc += cw
    y -= 16; pg = 1

    for i, a in enumerate(dados[:200]):
        if y < 55:
            c.showPage(); pg += 1; draw_frame(pg); y = H - 100
        c.setFillColorRGB(0.06, 0.09, 0.18 if i % 2 == 0 else 0.04)
        c.rect(40, y - 12, W - 80, 14, fill=1, stroke=0)
        lat = a.get('lat'); lng = a.get('lng'); acc = a.get('accuracy')
        gps = '—'
        if lat is not None and lng is not None:
            try:
                gps = f"{float(lat):.4f},{float(lng):.4f}"
                if acc: gps += f"±{int(float(acc))}m"
            except Exception:
                gps = f"{lat},{lng}"
        row = [
            (str(a.get('id', '')),           25),
            ((a.get('token_id') or '')[:12], 90),
            ((a.get('timestamp') or a.get('created_at') or '')[:16], 100),
            ((a.get('type') or '')[:14],     90),
            ((a.get('area') or '')[:12],     80),
            (gps[:18],                       100),
        ]
        c.setFillColorRGB(0.83, 0.86, 0.94); c.setFont("Helvetica", 7.5); xc = 45
        for txt, cw in row:
            c.drawString(xc, y - 8, txt); xc += cw
        resolved = a.get('resolved')
        c.setFillColorRGB(0.2, 0.85 if resolved else 0.3, 0.5 if resolved else 0.15)
        c.drawString(xc, y - 8, "Resolvido" if resolved else "Ativo")
        if a.get('maps_url'):
            c.linkURL(a['maps_url'], (295, y - 12, 395, y), relative=0)
        y -= 14

    c.save(); buf.seek(0)
    fname = f"condosafe24-{now.strftime('%Y%m%d-%H%M')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=False, download_name=fname)


# ════════════════════════════════════════════════════════════════════════════
#  EXPORTAÇÃO EXCEL (sem dados pessoais)
# ════════════════════════════════════════════════════════════════════════════

@app.route('/export.xlsx')
@require_central
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        try:
            dados = listar_alertas()
        except Exception:
            dados = alertas

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertas CONDO-SAFE24"

        hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hdr_fill  = PatternFill("solid", fgColor="0066FF")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borda     = Border(
            left=Side(style="thin", color="CBD5E1"),   right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),    bottom=Side(style="thin", color="CBD5E1"),
        )
        alt_fill = PatternFill("solid", fgColor="EFF6FF")

        # ⚠️ Sem "Chamador" e "Contato" — apenas token e dados operacionais
        cols = ["ID", "Token ID", "Data/Hora", "Tipo", "Área", "Prioridade",
                "Descrição", "Latitude", "Longitude", "Condomínio", "Status", "Maps"]
        for ci, col in enumerate(cols, 1):
            cl = ws.cell(row=1, column=ci, value=col)
            cl.font = hdr_font; cl.fill = hdr_fill
            cl.alignment = hdr_align; cl.border = borda

        ws.row_dimensions[1].height = 30
        larguras = [6, 16, 18, 18, 18, 10, 28, 12, 12, 18, 10, 35]
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        prioridade_labels = {1: 'Crítico', 2: 'Alto', 3: 'Médio'}
        for ri, a in enumerate(dados, 2):
            row_data = [
                a.get('id', ''),
                a.get('token_id', ''),
                a.get('timestamp') or a.get('created_at', ''),
                a.get('type', ''),
                a.get('area') or a.get('location', ''),
                prioridade_labels.get(a.get('prioridade', 2), 'Alto'),
                a.get('description', ''),
                a.get('lat', ''),
                a.get('lng', ''),
                a.get('client_name', ''),
                'Resolvido' if a.get('resolved') else 'Ativo',
                a.get('maps_url', ''),
            ]
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                cl = ws.cell(row=ri, column=ci, value=val)
                cl.border = borda
                cl.alignment = Alignment(vertical="center", wrap_text=False)
                if fill: cl.fill = fill
                if ci == 11:
                    cl.font = Font(color="15803D" if a.get('resolved') else "DC2626",
                                   bold=True, name="Arial")

        now = datetime.now(TZ)
        ws.append([])
        ws.append([f"Relatório gerado em: {now.strftime('%d/%m/%Y %H:%M:%S')} | "
                   "CONDO-SAFE24 v3.0 — Privacidade por Design | "
                   "SpyNet Tecnologia Forense | CNPJ 64.000.808/0001-51 | "
                   "⚠️ Dados de identidade protegidos no vault cifrado (LGPD)"])

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"condosafe24-alertas-{now.strftime('%Y%m%d-%H%M')}.xlsx"
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=fname)
    except Exception:
        log.exception("Erro ao gerar Excel")
        return jsonify({'ok': False, 'error': 'Erro interno ao gerar Excel.'}), 500


# ════════════════════════════════════════════════════════════════════════════
#  Z-API WHATSAPP
# ════════════════════════════════════════════════════════════════════════════

def _enviar_whatsapp(mensagem: str):
    instance = os.environ.get('ZAPI_INSTANCE', '').strip()
    token    = os.environ.get('ZAPI_TOKEN', '').strip()
    phone    = os.environ.get('ZAPI_PHONE', '').strip()
    if not all([instance, token, phone]):
        return
    try:
        import urllib.request, json as _json
        url     = f"https://api.z-api.io/instances/{instance}/token/{token}/send-text"
        payload = _json.dumps({"phone": phone, "message": mensagem}).encode()
        req     = urllib.request.Request(url, data=payload,
                    headers={"Content-Type": "application/json"}, method="POST")
        urllib.request.urlopen(req, timeout=5)
        log.info(f"WhatsApp enviado para {phone}")
    except Exception as e:
        log.warning(f"Falha ao enviar WhatsApp: {e}")


# ════════════════════════════════════════════════════════════════════════════
#  ERROR HANDLERS
# ════════════════════════════════════════════════════════════════════════════

@app.errorhandler(404)
def e404(e):
    if request.is_json:
        return jsonify({'ok': False, 'error': 'Não encontrado.'}), 404
    return render_template('forbidden.html', msg='Página não encontrada.'), 404

@app.errorhandler(500)
def e500(e):
    log.exception("Erro 500")
    if request.is_json:
        return jsonify({'ok': False, 'error': 'Erro interno.'}), 500
    return render_template('forbidden.html', msg='Erro interno. Tente novamente.'), 500


# ════════════════════════════════════════════════════════════════════════════
#  INICIALIZAÇÃO
# ════════════════════════════════════════════════════════════════════════════

from db import (
    init_db, salvar_alerta, listar_alertas, resolver_alerta_db,
    limpar_alertas_db, stats_alertas, audit, listar_audit,
    revelar_identidade, expirar_vault, gerar_token_id
)

try:
    init_db()
    expirar_vault()   # limpa vault expirado no startup
    log.info("CONDO-SAFE24 v3.0 — Banco inicializado. Vault limpo.")
except Exception as e:
    log.warning(f"Banco SQLite indisponível, usando memória: {e}")

if __name__ == '__main__':
    port  = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
